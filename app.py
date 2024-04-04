from flask import Flask, request, render_template, jsonify
import cv2
import openpyxl

app = Flask(__name__)

# Initialize the camera URLs
camera_urls = {
    "doctor1": 'http://192.168.161.130:4747/video',
    "doctor2": 'http://192.168.161.3:4747/video',
    "doctor3": 'http://192.168.161.218:4747/video',
}

# Load the Haar Cascade face detector
face_cascade = cv2.CascadeClassifier(cv2.data.haarcascades + 'haarcascade_frontalface_default.xml')

# Global variables to store patient data for different doctors
doctor_patients = {
    "doctor1": [],
    "doctor2": [],
    "doctor3": [],
}

# Global variable to store patient data
patients = []

# Function to detect and return a face from the camera feed
def detect_and_return_face(camera_url):
    # Initialize the camera
    cap = cv2.VideoCapture(camera_url)
    if not cap.isOpened():
        return False

    while True:
        # Read a frame from the camera
        ret, frame = cap.read()
        if not ret:
            return False

        # Convert the frame to grayscale for face detection
        gray = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)

        # Detect faces in the frame
        faces = face_cascade.detectMultiScale(gray, scaleFactor=1.3, minNeighbors=5, minSize=(50, 50))

        # Check if at least one face is detected
        if len(faces) > 0:
            cap.release()
            cv2.destroyAllWindows()
            return True

        # Exit the loop when the 'q' key is pressed
        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    # Release the camera and close all OpenCV windows if no face is detected
    cap.release()
    cv2.destroyAllWindows()
    return False

# Function to create an Excel sheet if it doesn't exist
def create_excel_sheet(filename, sheetname):
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = sheetname
    labels = ["Name", "Age", "Sex", "Phone Number", "Email", "Doctor"]
    sheet.append(labels)
    workbook.save(filename)

# Function to insert patient information into an Excel sheet
def insert_patient_info(filename, sheetname, data):
    try:
        workbook = openpyxl.load_workbook(filename)
    except FileNotFoundError:
        create_excel_sheet(filename, sheetname)
        workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    sheet.append(data)
    workbook.save(filename)

# Function to get patient details from the user
def patient_details():
    name = request.form['name']
    age = request.form['age']
    sex = request.form['gender']
    p_no = request.form['phone']
    e_mail = request.form['email']
    details = [name, age, sex, p_no, e_mail]
    return details

# Function to compare the number of patients in the Excel sheets
def compare_patient_counts():
    doctor1_count = len(doctor_patients["doctor1"])
    doctor2_count = len(doctor_patients["doctor2"])
    doctor3_count = len(doctor_patients["doctor3"])
    return {"doctor1": doctor1_count, "doctor2": doctor2_count, "doctor3": doctor3_count}

# Function to assign a patient to the doctor with the fewest patients
def assign_patient_to_doctor(details):
    patient_counts = compare_patient_counts()
    min_count_doctor = min(patient_counts, key=patient_counts.get)

    doctor_patients[min_count_doctor].append(details)

    return min_count_doctor

# Function to assign a patient to the doctor with the fewest patients overall
def assign_patient_to_doctor_overall(details):
    doctors = list(camera_urls.keys())
    doctor_counts = {doctor: len([p for p in patients if p["doctor"] == doctor]) for doctor in doctors}
    min_count_doctor = min(doctor_counts, key=doctor_counts.get)

    details.append(min_count_doctor)
    insert_patient_info(f"{min_count_doctor}.xlsx", "patients", details)
    patients.append({"doctor": min_count_doctor})

    return min_count_doctor

# Function to automatically detect available doctors and assign the patient
def detect_and_assign_doctor(details):
    available_doctors = []

    for doctor, camera_url in camera_urls.items():
        is_doctor_available = detect_and_return_face(camera_url)
        if is_doctor_available:
            available_doctors.append(doctor)

    if available_doctors:
        return assign_patient_to_doctor(details)
    else:
        return assign_patient_to_doctor_overall(details)  # Always assign to the doctor with the least patients overall

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'POST':
        details = patient_details()
        selected_doctor = request.form['doctor']

        if selected_doctor == "None":
            # Automatically detect and assign a doctor
            selected_doctor = detect_and_assign_doctor(details)
            if selected_doctor is None:
                return jsonify({"success": False, "message": "No doctor is available!"})

        # Assign the patient to the selected doctor
        insert_patient_info(f"{selected_doctor}.xlsx", "patients", details)
        return jsonify({"success": True, "message": f"Appointment booked with {selected_doctor}!"})

    return render_template('index.html')

if __name__ == '__main__':
    app.run(debug=True)